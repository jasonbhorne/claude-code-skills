"""Build the Travel Expense Tracker.xlsx from scratch.

Run once to create the workbook. Refuses to overwrite an existing file; pass
--force to rebuild from scratch (which wipes any logged trips).
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

import _lib

TRACKER_PATH = _lib.TRACKER_PATH
FORM_URL = _lib.FORM_URL
PERSONAL = _lib.PERSONAL
RATES = _lib.RATES

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def build_reference(wb):
    ws = wb.create_sheet("Reference")
    ws["A1"] = "Static Personal Fields"
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:B1")

    rows = [
        ("Street", PERSONAL["street"]),
        ("City", PERSONAL["city"]),
        ("State", PERSONAL["state"]),
        ("Zip", PERSONAL["zip"]),
        ("School/Building", PERSONAL["school"]),
        ("Signature", PERSONAL["signature"]),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    ws["A9"] = "Rates"
    ws["A9"].font = SECTION_FONT
    ws["A9"].fill = SECTION_FILL
    ws.merge_cells("A9:B9")

    rate_rows = [
        ("Mileage rate ($/mile)", RATES["mileage"]),
        ("Breakfast", RATES["breakfast"]),
        ("Lunch", RATES["lunch"]),
        ("Dinner", RATES["dinner"]),
    ]
    for i, (k, v) in enumerate(rate_rows, start=10):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=v)
        if i == 10:
            c.number_format = '"$"0.000'
        else:
            c.number_format = '"$"0.00'

    ws["A15"] = "Eligibility Rules"
    ws["A15"].font = SECTION_FONT
    ws["A15"].fill = SECTION_FILL
    ws.merge_cells("A15:B15")

    rules = [
        ("Breakfast", "Departure before 6:30 AM"),
        ("Lunch", "Departure before 11:00 AM"),
        ("Dinner", "Return after 7:00 PM"),
    ]
    for i, (k, v) in enumerate(rules, start=16):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    ws["A20"] = "Form URL"
    ws["A20"].font = SECTION_FONT
    ws["A20"].fill = SECTION_FILL
    ws.merge_cells("A20:B20")
    ws["A21"] = FORM_URL
    ws["A21"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[21].height = 60

    autofit(ws, {"A": 26, "B": 60})

    return ws


def add_named_ranges(wb):
    defs = {
        "MileageRate": "Reference!$B$10",
        "BreakfastRate": "Reference!$B$11",
        "LunchRate": "Reference!$B$12",
        "DinnerRate": "Reference!$B$13",
        "Street": "Reference!$B$2",
        "City": "Reference!$B$3",
        "State": "Reference!$B$4",
        "Zip": "Reference!$B$5",
        "School": "Reference!$B$6",
        "Signature": "Reference!$B$7",
    }
    for name, ref in defs.items():
        wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


TRIPS_HEADERS = [
    "Trip ID",
    "Purpose",
    "City/State",
    "Depart",
    "Return",
    "# Days",
    "Miles",
    "Mileage $",
    "Travel/Parking $",
    "Lodging $",
    "Other $",
    "Meals $",
    "Grand Total",
    "Account #",
    "Status",
    "Submission Date",
    "Notes",
]
TRIPS_COL_WIDTHS = {
    "A": 18, "B": 30, "C": 22, "D": 12, "E": 12, "F": 8,
    "G": 8, "H": 12, "I": 16, "J": 14, "K": 12, "L": 12,
    "M": 14, "N": 16, "O": 13, "P": 16, "Q": 30,
}


def build_trips(wb):
    ws = wb.create_sheet("Trips")
    for c, h in enumerate(TRIPS_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(TRIPS_HEADERS))
    ws.freeze_panes = "A2"
    autofit(ws, TRIPS_COL_WIDTHS)

    seed = {
        "A": "TDASC-2026-04",
        "B": "TDASC (Attendance Conference)",
        "C": "Murfreesboro, TN",
        "D": date(2026, 4, 7),
        "E": date(2026, 4, 10),
        "F": "=E2-D2+1",
        "G": 506,
        "H": "=G2*MileageRate",
        "I": 0,
        "J": 0,
        "K": 0,
        "L": '=SUMIFS(Days!$I:$I,Days!$A:$A,A2)',
        "M": "=H2+I2+J2+K2+L2",
        "N": "",
        "O": "Submitted",
        "P": date(2026, 4, 26),
        "Q": "Submitted via Microsoft Forms 2026-04-26",
    }
    for col, val in seed.items():
        ws[f"{col}2"] = val

    money_cols = ["H", "I", "J", "K", "L", "M"]
    date_cols = ["D", "E", "P"]
    for r in range(2, 502):
        for col in money_cols:
            ws[f"{col}{r}"].number_format = '"$"#,##0.00'
        for col in date_cols:
            ws[f"{col}{r}"].number_format = "m/d/yyyy"
        ws[f"F{r}"].alignment = Alignment(horizontal="center")

    dv = DataValidation(
        type="list", formula1='"Pending,Submitted,Reimbursed"', allow_blank=True
    )
    dv.add("O2:O500")
    ws.add_data_validation(dv)

    return ws


DAYS_HEADERS = [
    "Trip ID",
    "Date",
    "Day #",
    "Depart Time",
    "Return Time",
    "B",
    "L",
    "D",
    "Day Total",
]
DAYS_COL_WIDTHS = {
    "A": 18, "B": 12, "C": 7, "D": 13, "E": 13,
    "F": 5, "G": 5, "H": 5, "I": 12,
}


def build_days(wb):
    ws = wb.create_sheet("Days")
    for c, h in enumerate(DAYS_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(DAYS_HEADERS))
    ws.freeze_panes = "A2"
    autofit(ws, DAYS_COL_WIDTHS)

    tdasc_rows = [
        ("TDASC-2026-04", date(2026, 4, 7), 1, 0, 0, 1),
        ("TDASC-2026-04", date(2026, 4, 8), 2, 0, 0, 1),
        ("TDASC-2026-04", date(2026, 4, 9), 3, 0, 0, 1),
        ("TDASC-2026-04", date(2026, 4, 10), 4, 0, 1, 0),
    ]
    for i, (tid, d, n, b, l, dn) in enumerate(tdasc_rows, start=2):
        ws.cell(row=i, column=1, value=tid)
        ws.cell(row=i, column=2, value=d).number_format = "m/d/yyyy"
        ws.cell(row=i, column=3, value=n).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=4, value=None)
        ws.cell(row=i, column=5, value=None)
        ws.cell(row=i, column=6, value=b).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=7, value=l).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=8, value=dn).alignment = Alignment(horizontal="center")
        ws.cell(
            row=i,
            column=9,
            value=f"=F{i}*BreakfastRate+G{i}*LunchRate+H{i}*DinnerRate",
        ).number_format = '"$"#,##0.00'

    for r in range(2, 1001):
        ws[f"B{r}"].number_format = "m/d/yyyy"
        ws[f"D{r}"].number_format = "h:mm AM/PM"
        ws[f"E{r}"].number_format = "h:mm AM/PM"
        ws[f"I{r}"].number_format = '"$"#,##0.00'

    return ws


def build_form_fill(wb):
    ws = wb.create_sheet("Form Fill", 0)
    ws["A1"] = "Pick a trip — values below match the Microsoft Form fields."
    ws["A1"].font = Font(italic=True, color="595959")
    ws.merge_cells("A1:C1")

    ws["A3"] = "Trip ID:"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = "TDASC-2026-04"
    ws["B3"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["B3"].font = Font(bold=True)
    ws["B3"].alignment = Alignment(horizontal="center")

    def vlookup(col_idx):
        return f"=IFERROR(VLOOKUP($B$3,Trips!$A:$Q,{col_idx},FALSE),\"\")"

    def day_meal_text(day_n):
        n = day_n
        return (
            f'=IF(SUMIFS(Days!$I:$I,Days!$A:$A,$B$3,Days!$C:$C,{n})=0,"(no meals)",'
            f'TRIM('
            f'IF(SUMIFS(Days!$F:$F,Days!$A:$A,$B$3,Days!$C:$C,{n})>0,"Breakfast ","")&'
            f'IF(SUMIFS(Days!$G:$G,Days!$A:$A,$B$3,Days!$C:$C,{n})>0,"Lunch ","")&'
            f'IF(SUMIFS(Days!$H:$H,Days!$A:$A,$B$3,Days!$C:$C,{n})>0,"Dinner","")'
            f'))'
        )

    rows = [
        ("1. Street Address", "=Street", None),
        ("2. City", "=City", None),
        ("3. State", "=State", None),
        ("4. Zip Code", "=Zip", None),
        ("5. School/Building", "=School", None),
        ("6. Purpose of Trip", vlookup(2), None),
        ("7. City/State Where Travel Occurred", vlookup(3), None),
        ("8. Departure Date", vlookup(4), "m/d/yyyy"),
        ("9. Return Date", vlookup(5), "m/d/yyyy"),
        ("10. Number of Miles in Personal Vehicle", vlookup(7), "0"),
        ("11. Mileage Reimbursement", vlookup(8), '"$"#,##0.00'),
        ("12. Air/Taxi/Uber/Lyft/Bus/Parking/Other", vlookup(9), '"$"#,##0.00'),
        ("13. Hotel/Lodging Reimbursement", vlookup(10), '"$"#,##0.00'),
        ("14. Other Expenses", vlookup(11), '"$"#,##0.00'),
        ("15. Day 1", day_meal_text(1), None),
        ("16. Add additional meals?", '=IF(IFERROR(VLOOKUP($B$3,Trips!$A:$F,6,FALSE),0)>1,"Yes","No")', None),
        ("17. Day 2", day_meal_text(2), None),
        ("18. Day 3", day_meal_text(3), None),
        ("19. Day 4", day_meal_text(4), None),
        ("20. Day 5", day_meal_text(5), None),
        ("21. Day 6", day_meal_text(6), None),
        ("22. Day 7", day_meal_text(7), None),
        ("23. Grand Total of Expenses", vlookup(13), '"$"#,##0.00'),
        ("25. Account Number", vlookup(14), None),
        ("26. Signature", "=Signature", None),
    ]
    start = 5
    for i, (label, formula, fmt) in enumerate(rows):
        r = start + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = Font(bold=True)
        a.alignment = Alignment(vertical="center")
        b = ws.cell(row=r, column=2, value=formula)
        if fmt:
            b.number_format = fmt
        b.alignment = Alignment(vertical="center")

    note_row = start + len(rows) + 1
    ws.cell(row=note_row, column=1, value="Form URL:").font = Font(bold=True)
    ws.cell(row=note_row, column=2, value="=Reference!A21").alignment = Alignment(
        wrap_text=True
    )
    ws.row_dimensions[note_row].height = 45

    autofit(ws, {"A": 42, "B": 50, "C": 4})
    ws.freeze_panes = "A5"
    return ws


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true",
                        help="Overwrite existing tracker (DESTROYS logged trips).")
    args = parser.parse_args()

    if os.path.exists(TRACKER_PATH) and not args.force:
        print(
            f"Refusing to overwrite existing tracker at {TRACKER_PATH}.\n"
            f"Pass --force to rebuild from scratch (this wipes logged trips).",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    build_reference(wb)
    add_named_ranges(wb)
    build_trips(wb)
    build_days(wb)
    build_form_fill(wb)

    wb._sheets = [
        wb["Form Fill"],
        wb["Trips"],
        wb["Days"],
        wb["Reference"],
    ]

    os.makedirs(os.path.dirname(TRACKER_PATH), exist_ok=True)
    wb.save(TRACKER_PATH)
    print(f"Wrote {TRACKER_PATH}")


if __name__ == "__main__":
    main()
