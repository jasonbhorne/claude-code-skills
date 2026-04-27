"""Shared helpers for the travel-expense skill.

Loads config from ~/.claude/.travel-expense-config.json (see config.example.json
in the skill root). Personal info, the tracker path, and the form URL all live
in that file — never in source.
"""
from __future__ import annotations

import json
import os
from datetime import date, datetime, time
from typing import Iterable

import openpyxl

CONFIG_PATH = os.path.expanduser("~/.claude/.travel-expense-config.json")


def load_config():
    if not os.path.exists(CONFIG_PATH):
        raise FileNotFoundError(
            f"Config not found at {CONFIG_PATH}. "
            f"Copy config.example.json from the skill directory to that path "
            f"and fill in your details."
        )
    with open(CONFIG_PATH) as f:
        return json.load(f)


CONFIG = load_config()
TRACKER_PATH = os.path.expanduser(CONFIG["tracker_path"])
FORM_URL = CONFIG["form_url"]
PERSONAL = CONFIG["personal"]
RATES = CONFIG["rates"]

MILEAGE_RATE = RATES["mileage"]
BREAKFAST = RATES["breakfast"]
LUNCH = RATES["lunch"]
DINNER = RATES["dinner"]


def _parse_cutoff(s):
    h, m = s.split(":")
    return time(int(h), int(m))


BREAKFAST_CUTOFF = _parse_cutoff(CONFIG["eligibility"]["breakfast_cutoff"])
LUNCH_CUTOFF = _parse_cutoff(CONFIG["eligibility"]["lunch_cutoff"])
DINNER_CUTOFF = _parse_cutoff(CONFIG["eligibility"]["dinner_cutoff"])


def load_workbook():
    if not os.path.exists(TRACKER_PATH):
        raise FileNotFoundError(
            f"Tracker not found at {TRACKER_PATH}. "
            f"Run scripts/build_tracker.py first."
        )
    return openpyxl.load_workbook(TRACKER_PATH)


def save_workbook(wb):
    wb.save(TRACKER_PATH)


def parse_date(s):
    if s is None or s == "":
        return None
    if isinstance(s, (date, datetime)):
        return s if isinstance(s, datetime) else datetime(s.year, s.month, s.day)
    return datetime.fromisoformat(s)


def parse_time(s):
    if s is None or s == "":
        return None
    if isinstance(s, time):
        return s
    if isinstance(s, datetime):
        return s.time()
    s = s.strip()
    for fmt in ("%H:%M", "%I:%M %p", "%I:%M%p", "%I%p", "%I %p"):
        try:
            return datetime.strptime(s.upper(), fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Could not parse time: {s!r}")


def eligibility_from_times(depart, returnt):
    """Return (B, L, D) tuple of 0/1 from depart and return times.

    Returns None for any rule whose input time is missing — caller fills in manually.
    """
    b = (1 if depart and depart <= BREAKFAST_CUTOFF else 0) if depart else None
    l = (1 if depart and depart <= LUNCH_CUTOFF else 0) if depart else None
    d = (1 if returnt and returnt >= DINNER_CUTOFF else 0) if returnt else None
    return b, l, d


def first_empty_row(ws, key_col=1, start=2):
    r = start
    while ws.cell(row=r, column=key_col).value is not None:
        r += 1
    return r


def find_trip_row(ws, trip_id):
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == trip_id:
            return r
    return None


def find_day_rows(ws, trip_id) -> Iterable[int]:
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == trip_id:
            rows.append(r)
    return rows


def trip_summary(ws, row):
    return {
        "trip_id": ws.cell(row=row, column=1).value,
        "purpose": ws.cell(row=row, column=2).value,
        "city_state": ws.cell(row=row, column=3).value,
        "depart": ws.cell(row=row, column=4).value,
        "return": ws.cell(row=row, column=5).value,
        "miles": ws.cell(row=row, column=7).value,
        "travel_parking": ws.cell(row=row, column=9).value,
        "lodging": ws.cell(row=row, column=10).value,
        "other": ws.cell(row=row, column=11).value,
        "account": ws.cell(row=row, column=14).value,
        "status": ws.cell(row=row, column=15).value,
        "submission_date": ws.cell(row=row, column=16).value,
        "notes": ws.cell(row=row, column=17).value,
    }


def fmt_date(d):
    if d is None:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%-m/%-d/%Y")
    return str(d)


def fmt_money(v):
    return f"${v:,.2f}" if v is not None else "$0.00"
