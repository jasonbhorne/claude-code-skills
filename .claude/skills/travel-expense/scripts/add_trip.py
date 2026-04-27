"""Append a trip to the tracker.

Usage:
  python add_trip.py < trip.json

Trip JSON schema:
{
  "trip_id": "TDASC-2026-04",
  "purpose": "TDASC (Attendance Conference)",
  "city_state": "Murfreesboro, TN",
  "depart_date": "2026-04-07",
  "return_date": "2026-04-10",
  "miles": 506,
  "travel_parking": 0,
  "lodging": 0,
  "other": 0,
  "account": "",
  "status": "Pending",
  "submission_date": null,
  "notes": "",
  "days": [
    {"day": 1, "date": "2026-04-07", "depart_time": "07:00", "return_time": "18:00",
     "B": 0, "L": 0, "D": 1},
    ...
  ]
}

Fields are optional — missing values default to 0/blank. If `days` is omitted, no
meal rows are written and meals total = $0. If a day has B/L/D set, those override
auto-eligibility from times. If only times are set, eligibility auto-fills.
"""
from __future__ import annotations

import json
import sys
from datetime import datetime
from openpyxl.styles import Alignment

import _lib


def fill_trip_row(ws, row, trip):
    ws.cell(row=row, column=1, value=trip["trip_id"])
    ws.cell(row=row, column=2, value=trip.get("purpose", ""))
    ws.cell(row=row, column=3, value=trip.get("city_state", ""))
    ws.cell(row=row, column=4, value=_lib.parse_date(trip.get("depart_date")))
    ws.cell(row=row, column=5, value=_lib.parse_date(trip.get("return_date")))
    ws.cell(row=row, column=6, value=f"=E{row}-D{row}+1")
    ws.cell(row=row, column=7, value=trip.get("miles", 0))
    ws.cell(row=row, column=8, value=f"=G{row}*MileageRate")
    ws.cell(row=row, column=9, value=trip.get("travel_parking", 0))
    ws.cell(row=row, column=10, value=trip.get("lodging", 0))
    ws.cell(row=row, column=11, value=trip.get("other", 0))
    ws.cell(
        row=row,
        column=12,
        value=f'=SUMIFS(Days!$I:$I,Days!$A:$A,A{row})',
    )
    ws.cell(row=row, column=13, value=f"=H{row}+I{row}+J{row}+K{row}+L{row}")
    ws.cell(row=row, column=14, value=trip.get("account", ""))
    ws.cell(row=row, column=15, value=trip.get("status", "Pending"))
    sd = trip.get("submission_date")
    ws.cell(row=row, column=16, value=_lib.parse_date(sd) if sd else None)
    ws.cell(row=row, column=17, value=trip.get("notes", ""))

    money_cols = [8, 9, 10, 11, 12, 13]
    for c in money_cols:
        ws.cell(row=row, column=c).number_format = '"$"#,##0.00'
    for c in [4, 5, 16]:
        ws.cell(row=row, column=c).number_format = "m/d/yyyy"
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")


def fill_day_row(ws, row, trip_id, day):
    ws.cell(row=row, column=1, value=trip_id)
    ws.cell(row=row, column=2, value=_lib.parse_date(day.get("date")))
    ws.cell(row=row, column=3, value=day.get("day"))

    depart_t = _lib.parse_time(day.get("depart_time"))
    return_t = _lib.parse_time(day.get("return_time"))
    ws.cell(row=row, column=4, value=depart_t)
    ws.cell(row=row, column=5, value=return_t)

    auto_b, auto_l, auto_d = _lib.eligibility_from_times(depart_t, return_t)
    b = day["B"] if "B" in day else (auto_b if auto_b is not None else 0)
    l = day["L"] if "L" in day else (auto_l if auto_l is not None else 0)
    d = day["D"] if "D" in day else (auto_d if auto_d is not None else 0)

    ws.cell(row=row, column=6, value=b)
    ws.cell(row=row, column=7, value=l)
    ws.cell(row=row, column=8, value=d)
    ws.cell(
        row=row,
        column=9,
        value=f"=F{row}*BreakfastRate+G{row}*LunchRate+H{row}*DinnerRate",
    )

    ws.cell(row=row, column=2).number_format = "m/d/yyyy"
    ws.cell(row=row, column=4).number_format = "h:mm AM/PM"
    ws.cell(row=row, column=5).number_format = "h:mm AM/PM"
    ws.cell(row=row, column=9).number_format = '"$"#,##0.00'
    for c in [3, 6, 7, 8]:
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")


def main():
    raw = sys.stdin.read()
    if not raw.strip():
        print("error: trip JSON required on stdin", file=sys.stderr)
        sys.exit(2)
    trip = json.loads(raw)

    if "trip_id" not in trip:
        print("error: trip_id is required", file=sys.stderr)
        sys.exit(2)

    wb = _lib.load_workbook()
    trips = wb["Trips"]
    days = wb["Days"]

    if _lib.find_trip_row(trips, trip["trip_id"]) is not None:
        print(
            f"error: trip {trip['trip_id']} already exists. "
            f"Use a unique Trip ID or update existing.",
            file=sys.stderr,
        )
        sys.exit(2)

    row = _lib.first_empty_row(trips)
    fill_trip_row(trips, row, trip)

    day_entries = trip.get("days", [])
    for day in day_entries:
        drow = _lib.first_empty_row(days)
        fill_day_row(days, drow, trip["trip_id"], day)

    _lib.save_workbook(wb)
    print(f"Added trip {trip['trip_id']} (Trips row {row}, {len(day_entries)} day rows)")


if __name__ == "__main__":
    main()
