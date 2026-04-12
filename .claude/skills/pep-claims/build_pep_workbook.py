#!/usr/bin/env python3
"""
Build PEP Claims Metrics workbook for Greeneville City Schools.

Usage:
  python build_pep_workbook.py <output_path> <prior_path_or_empty> <json_data_file>

json_data_file format:
{
  "pulled_at": "2026-04-12",
  "updated_as_of": "4/12/2026 3:30:02 AM",
  "types": {
    "Workers' Comp":       {"open":3, "incurred":656482, "top_open_count":2, "top_open_total":640137, "all_count":616, "all_total":2224442, "litigated":0,
                            "claims_by_year":{"2020":11,"2021":14,"2022":14,"2023":41,"2024":45,"2025":20},
                            "net_incurred_by_year":{"2016":73800,"2017":88300,"2018":24400,"2019":54100,"2020":132000,"2021":737800,"2022":87300,"2023":81100,"2024":166300,"2025":21500},
                            "claims_by_month":{"May 2025":6,"Jun 2025":2,"Jul 2025":2,"Aug 2025":3,"Sep 2025":6,"Oct 2025":1,"Nov 2025":0,"Dec 2025":2,"Jan 2026":1,"Feb 2026":3,"Mar 2026":2,"Apr 2026":0}},
    ...
  }
}
"""

import sys
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Style helpers
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(italic=True, size=10, color="666666")
YELLOW_FILL = PatternFill("solid", fgColor="FFF9C4")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL = PatternFill("solid", fgColor="FCE4EC")
CURRENCY = '$#,##0'
PCT = '0.0%'
THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Canonical order for insurance types
TYPE_ORDER = [
    "Workers' Comp",
    "Liability (All)",
    "General Liability",
    "Property",
    "Auto Physical Damage",
    "Auto Liability",
    "E&O",
    "Law Enforcement",
]

YEARS_10 = [str(y) for y in range(2016, 2026)]
YEARS_6 = [str(y) for y in range(2020, 2026)]


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN


def data_cell(ws, row, col, value=None, fmt=None, bold=False, center=True):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.border = THIN
    if center and col > 1:
        c.alignment = Alignment(horizontal="center")
    if fmt:
        c.number_format = fmt
    if bold:
        c.font = Font(bold=True)
    return c


def norm_int(v):
    if v is None:
        return 0
    if isinstance(v, str):
        return int(v.replace(',', '').replace('$', '') or 0)
    return int(v)


def build_summary(ws, data):
    ws['A1'] = "PEP Claims Metrics - Greeneville City Schools"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Data as of {data.get('updated_as_of', '')} | Pulled {data.get('pulled_at', '')}"
    ws['A2'].font = SUBTITLE_FONT

    headers = ["Insurance Type", "Open Claims", "Open Incurred Loss", "Top Open Claims",
               "Top Open $ Incurred", "Total Claims (All)", "Total $ Incurred (All)", "Litigated"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    for i, tname in enumerate(TYPE_ORDER):
        t = data['types'].get(tname, {})
        row = 5 + i
        data_cell(ws, row, 1, tname, center=False)
        data_cell(ws, row, 2, norm_int(t.get('open', 0)))
        data_cell(ws, row, 3, norm_int(t.get('incurred', 0)), CURRENCY)
        data_cell(ws, row, 4, norm_int(t.get('top_open_count', 0)))
        data_cell(ws, row, 5, norm_int(t.get('top_open_total', 0)), CURRENCY)
        data_cell(ws, row, 6, norm_int(t.get('all_count', 0)))
        data_cell(ws, row, 7, norm_int(t.get('all_total', 0)), CURRENCY)
        data_cell(ws, row, 8, norm_int(t.get('litigated', 0)))

    tr = 5 + len(TYPE_ORDER)
    data_cell(ws, tr, 1, "TOTAL", bold=True, center=False)
    for col in [2, 4, 6, 8]:
        data_cell(ws, tr, col, f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{tr-1})", bold=True)
    for col in [3, 5, 7]:
        data_cell(ws, tr, col, f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{tr-1})", CURRENCY, bold=True)

    ws.column_dimensions['A'].width = 22
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18


def build_net_incurred(ws, data):
    ws['A1'] = "Net Incurred by Policy Year (2016-2025)"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = "Source: PEP Claims Financial Metrics"
    ws['A2'].font = SUBTITLE_FONT

    hdrs = ["Insurance Type"] + YEARS_10 + ["Total"]
    for col, h in enumerate(hdrs, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(hdrs))

    for i, tname in enumerate(TYPE_ORDER):
        t = data['types'].get(tname, {})
        year_vals = t.get('net_incurred_by_year', {})
        row = 5 + i
        data_cell(ws, row, 1, tname, center=False)
        for j, yr in enumerate(YEARS_10):
            data_cell(ws, row, j + 2, norm_int(year_vals.get(yr, 0)), CURRENCY)
        data_cell(ws, row, 12, f"=SUM(B{row}:K{row})", CURRENCY, bold=True)

    tr = 5 + len(TYPE_ORDER)
    data_cell(ws, tr, 1, "TOTAL ALL LINES", bold=True, center=False)
    ws.cell(row=tr, column=1).fill = YELLOW_FILL
    for col in range(2, 13):
        c = data_cell(ws, tr, col, f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{tr-1})", CURRENCY, bold=True)
        c.fill = YELLOW_FILL

    ws.column_dimensions['A'].width = 22
    for col in range(2, 13):
        ws.column_dimensions[get_column_letter(col)].width = 13


def build_yoy(ws):
    ws['A1'] = "Year-over-Year Change in Net Incurred"
    ws['A1'].font = TITLE_FONT

    periods = [f"{YEARS_10[i]}-{YEARS_10[i+1][-2:]}" for i in range(9)]
    hdrs = ["Insurance Type"] + periods
    for col, h in enumerate(hdrs, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(hdrs))

    ref = "'Net Incurred (10-Year)'"
    for i, tname in enumerate(TYPE_ORDER):
        row = 5 + i
        data_row = 5 + i
        data_cell(ws, row, 1, tname, center=False)
        for j in range(9):
            cur = get_column_letter(j + 3)
            prev = get_column_letter(j + 2)
            data_cell(ws, row, j + 2, f"={ref}!{cur}{data_row}-{ref}!{prev}{data_row}", CURRENCY)

    tr = 5 + len(TYPE_ORDER)
    data_cell(ws, tr, 1, "TOTAL ALL LINES", bold=True, center=False)
    for col in range(2, 11):
        c = data_cell(ws, tr, col, f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{tr-1})", CURRENCY, bold=True)
        c.fill = YELLOW_FILL

    # Percentage section
    pct_start = tr + 3
    ws.cell(row=pct_start - 1, column=1, value="Percentage Change Year-over-Year").font = Font(bold=True, size=12, color="2F5496")
    for col, h in enumerate(hdrs, 1):
        ws.cell(row=pct_start, column=col, value=h)
    style_header_row(ws, pct_start, len(hdrs))

    for i, tname in enumerate(TYPE_ORDER):
        row = pct_start + 1 + i
        data_row = 5 + i
        data_cell(ws, row, 1, tname, center=False)
        for j in range(9):
            cur = get_column_letter(j + 3)
            prev = get_column_letter(j + 2)
            data_cell(ws, row, j + 2,
                      f"=IF({ref}!{prev}{data_row}=0,0,({ref}!{cur}{data_row}-{ref}!{prev}{data_row})/{ref}!{prev}{data_row})",
                      PCT)

    ws.column_dimensions['A'].width = 22
    for col in range(2, 11):
        ws.column_dimensions[get_column_letter(col)].width = 13


def build_claims_by_year(ws, data):
    ws['A1'] = "Claims Count by Policy Start Year (2020-2025)"
    ws['A1'].font = TITLE_FONT

    hdrs = ["Insurance Type"] + YEARS_6 + ["Total"]
    for col, h in enumerate(hdrs, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(hdrs))

    for i, tname in enumerate(TYPE_ORDER):
        t = data['types'].get(tname, {})
        year_vals = t.get('claims_by_year', {})
        row = 5 + i
        data_cell(ws, row, 1, tname, center=False)
        for j, yr in enumerate(YEARS_6):
            data_cell(ws, row, j + 2, norm_int(year_vals.get(yr, 0)))
        data_cell(ws, row, 8, f"=SUM(B{row}:G{row})", bold=True)

    tr = 5 + len(TYPE_ORDER)
    data_cell(ws, tr, 1, "TOTAL", bold=True, center=False)
    for col in range(2, 9):
        data_cell(ws, tr, col, f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{tr-1})", bold=True)

    ws.column_dimensions['A'].width = 22
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 10


def build_claims_by_month(ws, data):
    ws['A1'] = "Claims by Loss Month (Last 12 Months)"
    ws['A1'].font = TITLE_FONT

    # Derive months from the first type that has them
    months = []
    for tname in TYPE_ORDER:
        m = data['types'].get(tname, {}).get('claims_by_month')
        if m:
            months = list(m.keys())
            break
    if not months:
        months = ["Month"] * 12

    hdrs = ["Insurance Type"] + months + ["Total"]
    for col, h in enumerate(hdrs, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(hdrs))

    for i, tname in enumerate(TYPE_ORDER):
        t = data['types'].get(tname, {})
        m_vals = t.get('claims_by_month', {})
        row = 4 + i
        data_cell(ws, row, 1, tname, center=False)
        for j, mname in enumerate(months):
            data_cell(ws, row, j + 2, norm_int(m_vals.get(mname, 0)))
        data_cell(ws, row, len(months) + 2, f"=SUM(B{row}:{get_column_letter(len(months)+1)}{row})", bold=True)

    tr = 4 + len(TYPE_ORDER)
    data_cell(ws, tr, 1, "TOTAL", bold=True, center=False)
    for col in range(2, len(months) + 3):
        data_cell(ws, tr, col, f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{tr-1})", bold=True)

    ws.column_dimensions['A'].width = 22
    for col in range(2, len(months) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 11


def build_combined(ws):
    ws['A1'] = "All Insurance Lines Combined - 10 Year Trend"
    ws['A1'].font = TITLE_FONT

    hdrs = ["Year", "Workers' Comp", "Liability", "Property", "Auto Phys Damage", "Auto Liability", "E&O", "Total All Lines", "YoY Change"]
    for col, h in enumerate(hdrs, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(hdrs))

    ref = "'Net Incurred (10-Year)'"
    # Rows in Net Incurred sheet: WC=5, Liability=6, GenLiab=7(skip), Property=8, APD=9, AutoLiab=10, E&O=11
    for i, year in enumerate(YEARS_10):
        row = 5 + i
        col_letter = get_column_letter(i + 2)
        data_cell(ws, row, 1, int(year), '0', center=False)
        data_cell(ws, row, 2, f"={ref}!{col_letter}5", CURRENCY)
        data_cell(ws, row, 3, f"={ref}!{col_letter}6", CURRENCY)
        data_cell(ws, row, 4, f"={ref}!{col_letter}8", CURRENCY)
        data_cell(ws, row, 5, f"={ref}!{col_letter}9", CURRENCY)
        data_cell(ws, row, 6, f"={ref}!{col_letter}10", CURRENCY)
        data_cell(ws, row, 7, f"={ref}!{col_letter}11", CURRENCY)
        data_cell(ws, row, 8, f"=SUM(B{row}:G{row})", CURRENCY, bold=True)
        if i == 0:
            data_cell(ws, row, 9, 0, PCT)
        else:
            data_cell(ws, row, 9, f"=IF(H{row-1}=0,0,(H{row}-H{row-1})/H{row-1})", PCT)

    tr = 15
    data_cell(ws, tr, 1, "10-Yr Average", bold=True, center=False)
    for col in range(2, 9):
        c = data_cell(ws, tr, col, f"=AVERAGE({get_column_letter(col)}5:{get_column_letter(col)}14)", CURRENCY, bold=True)
        c.fill = YELLOW_FILL

    ws.column_dimensions['A'].width = 14
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 16


def load_prior_summary(prior_path):
    """Read the Summary tab from a prior workbook, returning a dict keyed by insurance type."""
    if not prior_path or not Path(prior_path).exists():
        return None
    try:
        pwb = load_workbook(prior_path, data_only=True)
        ws = pwb['Summary']
    except Exception:
        return None

    prior = {}
    for i in range(len(TYPE_ORDER)):
        row = 5 + i
        tname = ws.cell(row=row, column=1).value
        if not tname:
            continue
        prior[tname.strip()] = {
            'open': norm_int(ws.cell(row=row, column=2).value),
            'incurred': norm_int(ws.cell(row=row, column=3).value),
            'top_open_count': norm_int(ws.cell(row=row, column=4).value),
            'top_open_total': norm_int(ws.cell(row=row, column=5).value),
            'all_count': norm_int(ws.cell(row=row, column=6).value),
            'all_total': norm_int(ws.cell(row=row, column=7).value),
        }
    return prior


def build_mom(ws, data, prior):
    ws['A1'] = "Month-over-Month Changes"
    ws['A1'].font = TITLE_FONT
    if prior is None:
        ws['A2'] = "No prior month workbook found - this is the baseline snapshot."
        ws['A2'].font = SUBTITLE_FONT
        return
    ws['A2'] = "Comparison vs. most recent prior monthly pull."
    ws['A2'].font = SUBTITLE_FONT

    hdrs = ["Insurance Type", "Open Claims Δ", "Incurred Loss Δ", "Total Claims Δ", "Total $ Incurred Δ", "Flag"]
    for col, h in enumerate(hdrs, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(hdrs))

    for i, tname in enumerate(TYPE_ORDER):
        cur = data['types'].get(tname, {})
        prev = prior.get(tname, {})
        row = 5 + i

        open_delta = norm_int(cur.get('open', 0)) - prev.get('open', 0)
        incurred_delta = norm_int(cur.get('incurred', 0)) - prev.get('incurred', 0)
        all_count_delta = norm_int(cur.get('all_count', 0)) - prev.get('all_count', 0)
        all_total_delta = norm_int(cur.get('all_total', 0)) - prev.get('all_total', 0)

        data_cell(ws, row, 1, tname, center=False)
        c_o = data_cell(ws, row, 2, open_delta, '+#,##0;-#,##0;0')
        c_i = data_cell(ws, row, 3, incurred_delta, '+$#,##0;-$#,##0;$0')
        c_c = data_cell(ws, row, 4, all_count_delta, '+#,##0;-#,##0;0')
        c_t = data_cell(ws, row, 5, all_total_delta, '+$#,##0;-$#,##0;$0')

        flags = []
        if open_delta > 0:
            flags.append(f"+{open_delta} new open claim(s)")
            c_o.fill = RED_FILL
        elif open_delta < 0:
            c_o.fill = GREEN_FILL
        if incurred_delta > 10000:
            flags.append(f"Incurred up >${incurred_delta:,.0f}")
            c_i.fill = RED_FILL
        elif incurred_delta < -10000:
            c_i.fill = GREEN_FILL
        if all_count_delta > 0:
            flags.append(f"{all_count_delta} new claim(s) filed")
        data_cell(ws, row, 6, "; ".join(flags) if flags else "no change", center=False)

    tr = 5 + len(TYPE_ORDER)
    data_cell(ws, tr, 1, "TOTAL", bold=True, center=False)
    for col in [2, 4]:
        data_cell(ws, tr, col, f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{tr-1})",
                  '+#,##0;-#,##0;0', bold=True)
    for col in [3, 5]:
        data_cell(ws, tr, col, f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{tr-1})",
                  '+$#,##0;-$#,##0;$0', bold=True)

    ws.column_dimensions['A'].width = 22
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18


def main():
    if len(sys.argv) < 4:
        print("Usage: build_pep_workbook.py <output_path> <prior_path_or_empty> <json_data_file>")
        sys.exit(1)

    output_path = sys.argv[1]
    prior_path = sys.argv[2] if sys.argv[2] else None
    data_file = sys.argv[3]

    data = json.loads(Path(data_file).read_text())
    prior = load_prior_summary(prior_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    build_summary(ws, data)
    build_net_incurred(wb.create_sheet("Net Incurred (10-Year)"), data)
    build_yoy(wb.create_sheet("YoY Analysis"))
    build_claims_by_year(wb.create_sheet("Claims Count by Year"), data)
    build_claims_by_month(wb.create_sheet("Claims by Month"), data)
    build_combined(wb.create_sheet("All Lines Combined"))
    build_mom(wb.create_sheet("MoM Changes"), data, prior)

    # Ensure output dir exists
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}")
    if prior:
        print(f"Compared against: {prior_path}")
    else:
        print("No prior month - baseline snapshot")


if __name__ == "__main__":
    main()
